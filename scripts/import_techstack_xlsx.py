"""Import 260622_FaSS_TechStack_With_Assignees xlsx → temp_techstack.json + fassTechStack.ts"""
import json
import subprocess
import sys
from collections import OrderedDict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path(r"c:\Users\user\Desktop\260622_FaSS_TechStack_With_Assignees_심지훈 수정.xlsx")

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

rows_out = []
current_phase = None
current_sprint = None
current_task = None
current_assignee = None

for r in range(2, ws.max_row + 1):
    phase = ws.cell(r, 1).value
    sprint = ws.cell(r, 2).value
    task = ws.cell(r, 3).value
    assignee = ws.cell(r, 4).value
    tech = ws.cell(r, 5).value
    desc = ws.cell(r, 6).value

    if phase:
        current_phase = str(phase).strip()
    if sprint:
        current_sprint = str(sprint).strip()
        current_assignee = str(assignee).strip() if assignee else ""
    if task:
        current_task = str(task).strip()
        if assignee:
            current_assignee = str(assignee).strip()

    if not tech:
        continue

    rows_out.append(
        {
            "단계": current_phase,
            "스프린트": current_sprint,
            "과제": current_task,
            "메인담당": current_assignee,
            "기술 스택": str(tech).strip(),
            "역할 설명": str(desc).strip() if desc else "",
        }
    )

tasks = OrderedDict()
for row in rows_out:
    key = (row["스프린트"], row["과제"])
    if key not in tasks:
        tasks[key] = {
            "phase": row["단계"],
            "sprint": row["스프린트"],
            "task": row["과제"],
            "assignee": row["메인담당"] or "",
            "techs": [],
        }
    if row["메인담당"] and not tasks[key]["assignee"]:
        tasks[key]["assignee"] = row["메인담당"]
    tasks[key]["techs"].append({"name": row["기술 스택"], "desc": row["역할 설명"]})

payload = {"source": XLSX.name, "rows": rows_out, "tasks": list(tasks.values())}
(ROOT / "temp_techstack.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {len(rows_out)} tech rows, {len(tasks)} tasks to temp_techstack.json")

subprocess.check_call([sys.executable, str(ROOT / "scripts" / "gen_techstack.py")], cwd=ROOT)
